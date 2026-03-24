import csv
import re
from collections import defaultdict
from decimal import Decimal

import pandas as pd


THRESHOLD = 40
COMP_TO_DATE_REQUIRED = "99999999"
NOTES_PREFIXES = ("BGGXHWP", "BGGXKWP")

ALLBOM_PARTNO_CANDIDATES = [
    "Product number",
    "Product Number",
    "Product no",
    "Product No",
    "Product No.",
    "Component/WC",
    "Component/WC ",
    "Compenent/WC",
    "Compenent/WC ",
]

ALLBOM_QTY_CANDIDATES = [
    "Amount",
    "Qty",
    "Quantity",
    "quantity",
    "amount",
    "Qty/Min",
    "Qty / Min",
    "Qty/ Min",
    "Qty/Min ",
]

PRODUCT_CODE_COL_CANDIDATES = [
    "Product code",
    "Product Code",
    "Product",
    "Product name",
    "Unit",
    "Unit code",
    "Product ID",
    "Product No",
    "Product no",
    "Project",
    "Project code",
]


def normalize_id(x):
    s = (x or "").strip()
    if s == "":
        return ""
    low = s.casefold().strip()
    if low in {"n/a", "none", "nan", "na", ".", "..", "...", "-", "--", "---", "\ufffd"}:
        return ""
    if re.fullmatch(r"[-\.\ufffd]+", s):
        return ""
    s = s.replace(",", "")
    if re.fullmatch(r"\d+\.(0+)", s):
        s = s.split(".")[0]
    if re.fullmatch(r"\d+(\.\d+)?[eE][+-]?\d+", s):
        try:
            s = str(int(Decimal(s)))
        except Exception:
            pass
    return s.strip()


def to_float_safe(x):
    try:
        s = (x or "").strip()
        if s == "":
            return 0.0
        return float(s.replace(",", ""))
    except Exception:
        return 0.0


def to_int_safe(x):
    try:
        s = (x or "").strip()
        if s == "":
            return None
        s = s.replace(",", "")
        m = re.search(r"-?\d+(\.\d+)?", s)
        if not m:
            return None
        return int(float(m.group(0)))
    except Exception:
        return None


def normalize_makebuy(v):
    v = (v or "").strip().casefold()
    if ("purch" in v) or (v in {"buy", "bought", "purchase"}):
        return "purchased"
    if ("manuf" in v) or ("mfg" in v) or ("make" in v):
        return "manufacturing"
    return v


def normalize_acq(v):
    v = (v or "").strip().casefold()
    if "purch" in v:
        return "purchased"
    if ("manuf" in v) or ("mfg" in v) or ("make" in v):
        return "manufacturing"
    return v


def prefer_nonempty(old, new):
    old = (old or "").strip()
    new = (new or "").strip()
    return new if (new and not old) else old


def normalize_comp_to_date(x):
    s = (x or "").strip().replace(",", "")
    if re.fullmatch(r"\d+\.(0+)", s):
        s = s.split(".")[0]
    return s


def safe_excel_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "_", name or "").strip()
    name = name[:31] if len(name) > 31 else name
    return name or "Sheet"


def clean_allbom_sheet_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()

    if df.shape[1] > 0:
        first_col = df.iloc[:, 0]
        if first_col.isna().all() or first_col.astype(str).str.strip().str.replace("nan", "").eq("").all():
            df = df.iloc[:, 1:]

    header_row_idx = None
    candidates_cf = {c.strip().casefold() for c in ALLBOM_PARTNO_CANDIDATES}

    for i in range(min(60, len(df))):
        row_vals = df.iloc[i].tolist()
        row_strs = [
            str(x).strip().casefold()
            for x in row_vals
            if str(x).strip() and str(x).strip().casefold() != "nan"
        ]
        if not row_strs:
            continue
        if any(s in candidates_cf for s in row_strs):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise Exception(
            "Could not find AllBOM header row. Expected a column like 'Product number' or 'Component/WC'."
        )

    header = df.iloc[header_row_idx].tolist()
    df = df.iloc[header_row_idx + 1 :].copy()
    df.columns = [
        ""
        if (h is None or (isinstance(h, float) and pd.isna(h)))
        else str(h).strip()
        for h in header
    ]

    df = df.dropna(how="all")

    drop_cols = []
    for c in df.columns:
        c_str = str(c).strip()
        c_cf = c_str.casefold()
        if c_str == "" or c_cf == "nan" or c_cf.startswith("unnamed"):
            drop_cols.append(c)
    if drop_cols:
        df = df.drop(columns=drop_cols)

    col_map_cf = {str(c).strip().casefold(): c for c in df.columns}
    found_part_col = None
    for cand in ALLBOM_PARTNO_CANDIDATES:
        key = cand.strip().casefold()
        if key in col_map_cf:
            found_part_col = col_map_cf[key]
            break
    if found_part_col is None:
        raise Exception("Header found, but could not locate the part-number column after cleaning.")
    if found_part_col != "Product number":
        df = df.rename(columns={found_part_col: "Product number"})

    return df


def find_product_code_column(df: pd.DataFrame):
    cols = list(df.columns)
    lower_map = {str(c).strip().casefold(): c for c in cols}

    for cand in PRODUCT_CODE_COL_CANDIDATES:
        key = cand.strip().casefold()
        if key in lower_map:
            return lower_map[key]

    for c in cols:
        cl = str(c).strip().casefold()
        if "product" in cl and "code" in cl:
            return c
    return None


def build_part_to_productcode_map(allbom_df: pd.DataFrame) -> dict:
    part_to_code = {}
    if "Product number" not in allbom_df.columns:
        return part_to_code
    code_col = find_product_code_column(allbom_df)
    if code_col is None:
        return part_to_code

    for _, row in allbom_df.iterrows():
        pn = normalize_id(str(row.get("Product number", "") or ""))
        if not pn:
            continue
        code = str(row.get(code_col, "") or "").strip()
        if pn not in part_to_code and code:
            part_to_code[pn] = code
    return part_to_code


def load_component_db(path="component_db.csv"):
    edges = defaultdict(list)
    parents_set = set()
    comp_details = {}
    product_comp_to_date = {}

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        required = [
            "Product no",
            "Component no",
            "Component name",
            "Component description",
            "Make/buy",
            "Quantity",
            "Comp To date",
        ]
        missing = [c for c in required if c not in fields]
        if missing:
            raise Exception(f"{path} missing columns: {missing}. Found: {fields}")

        for r in reader:
            parent = normalize_id(r.get("Product no"))
            child = normalize_id(r.get("Component no"))
            if not parent or not child:
                continue

            edge_qty = to_float_safe(r.get("Quantity"))
            if edge_qty <= 0:
                edge_qty = 1.0

            parents_set.add(parent)
            edges[parent].append((child, edge_qty))

            ctd_parent = normalize_comp_to_date(r.get("Comp To date"))
            if ctd_parent:
                oldp = product_comp_to_date.get(parent, "")
                if oldp == COMP_TO_DATE_REQUIRED:
                    product_comp_to_date[parent] = oldp
                elif ctd_parent == COMP_TO_DATE_REQUIRED:
                    product_comp_to_date[parent] = ctd_parent
                elif oldp == "":
                    product_comp_to_date[parent] = ctd_parent

            existing = comp_details.get(child, {})
            mb_new = normalize_makebuy(r.get("Make/buy"))
            mb_old = normalize_makebuy(existing.get("Make/buy"))
            mb_final = "purchased" if (mb_new == "purchased" or mb_old == "purchased") else (mb_old or mb_new)

            ctd_new = normalize_comp_to_date(r.get("Comp To date"))
            ctd_old = normalize_comp_to_date(existing.get("Comp To date", ""))
            ctd_final = (
                COMP_TO_DATE_REQUIRED
                if (ctd_new == COMP_TO_DATE_REQUIRED or ctd_old == COMP_TO_DATE_REQUIRED)
                else (ctd_old or ctd_new)
            )

            comp_details[child] = {
                "Component no": child,
                "Component name": prefer_nonempty(existing.get("Component name"), r.get("Component name")),
                "Component description": prefer_nonempty(
                    existing.get("Component description"),
                    r.get("Component description"),
                ),
                "Make/buy": mb_final,
                "Comp To date": ctd_final,
            }

    return edges, parents_set, comp_details, product_comp_to_date


def load_leadtime(path="leadtime.csv"):
    lead_map = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        required = ["Item no", "Item name", "Lead time"]
        missing = [c for c in required if c not in fields]
        if missing:
            raise Exception(f"{path} missing columns: {missing}. Found: {fields}")

        acq_col = next((c for c in ("Acquisition code", "Acquisition", "Acq code", "Acq. code") if c in fields), None)
        supplier_col = next(
            (
                c
                for c in (
                    "Supplier",
                    "Supplier name",
                    "Supplier Name",
                    "Vendor",
                    "Vendor name",
                    "Vendor Name",
                    "Supplier/Vendor",
                    "Supplier/Vendor name",
                    "Manufacturer",
                    "Manufacturer name",
                    "Vendor No",
                    "Supplier No",
                )
                if c in fields
            ),
            None,
        )
        safety_col = next(
            (
                c
                for c in (
                    "Safety stock",
                    "Safety Stock",
                    "SafetyStock",
                    "Safety stock qty",
                    "Safety Stock Qty",
                    "Safety stock quantity",
                    "Safety",
                    "Safetystock",
                    "SS",
                    "Min stock",
                    "Minimum stock",
                    "Min qty",
                    "Minimum qty",
                    "Reorder point",
                    "ROP",
                )
                if c in fields
            ),
            None,
        )

        for r in reader:
            item = normalize_id(r.get("Item no"))
            if not item:
                continue
            lead_map[item] = {
                "Item no": item,
                "Item name": (r.get("Item name") or "").strip(),
                "Lead time": to_int_safe(r.get("Lead time")),
                "Acq": normalize_acq(r.get(acq_col)) if acq_col else "",
                "Supplier": (r.get(supplier_col) or "").strip() if supplier_col else "",
                "Safety stock": to_float_safe(r.get(safety_col)) if safety_col else "",
            }
    return lead_map


def process_allbom_df(allbom_df, edges, parents_set, comp_details, product_comp_to_date, lead_map):
    cols = list(allbom_df.columns)
    if "Product number" not in cols:
        raise Exception(f"AllBOM sheet missing 'Product number'. Found: {cols}")

    qty_col = next((c for c in ALLBOM_QTY_CANDIDATES if c in cols), None)
    part_to_productcode = build_part_to_productcode_map(allbom_df)

    allbom_qty = defaultdict(float)
    for _, row in allbom_df.iterrows():
        pn = normalize_id(str(row.get("Product number", "") or ""))
        if not pn:
            continue
        qty = to_float_safe(str(row.get(qty_col, "") or "")) if qty_col else 1.0
        if qty <= 0:
            continue
        allbom_qty[pn] += qty

    start_kits = []
    direct_parts_qty = defaultdict(float)
    for pn, qty in allbom_qty.items():
        if pn in parents_set:
            start_kits.append((pn, qty))
        else:
            direct_parts_qty[pn] += qty

    leaf_qty = defaultdict(float)

    def dfs(parent, parent_qty, depth, path):
        if depth > 80 or parent in path:
            return
        children = edges.get(parent, [])
        if not children:
            leaf_qty[parent] += parent_qty
            return
        for child, edge_qty in children:
            child_total_qty = parent_qty * edge_qty
            if child in parents_set:
                dfs(child, child_total_qty, depth + 1, path | {parent})
            else:
                leaf_qty[child] += child_total_qty

    for kit, kit_qty in start_kits:
        dfs(kit, kit_qty, 0, set())

    for part, qty in direct_parts_qty.items():
        leaf_qty[part] += qty

    def is_purchased(part_no):
        lt = lead_map.get(part_no)
        if lt and lt.get("Acq"):
            return lt["Acq"] == "purchased"
        return normalize_makebuy(comp_details.get(part_no, {}).get("Make/buy", "")) == "purchased"

    def get_comp_to_date_for_part(part_no):
        ctd = normalize_comp_to_date(comp_details.get(part_no, {}).get("Comp To date", ""))
        if ctd:
            return ctd
        return normalize_comp_to_date(product_comp_to_date.get(part_no, ""))

    amo_rows = []
    for part, total_qty in leaf_qty.items():
        if total_qty <= 0:
            continue
        ltrow = lead_map.get(part)
        if not ltrow or ltrow.get("Lead time") is None:
            continue
        if not is_purchased(part):
            continue
        if ltrow["Lead time"] <= THRESHOLD:
            continue

        ctd = get_comp_to_date_for_part(part)
        if ctd and ctd != COMP_TO_DATE_REQUIRED:
            continue

        notes_val = part_to_productcode.get(part, "") if str(part).startswith(NOTES_PREFIXES) else ""
        comp = comp_details.get(part, {})

        amo_rows.append(
            {
                "Part number": part,
                "Qty": round(total_qty, 3),
                "Lead time": ltrow["Lead time"],
                "Item name": ltrow.get("Item name", ""),
                "Supplier": ltrow.get("Supplier", ""),
                "Component name": comp.get("Component name", ""),
                "Component description": comp.get("Component description", ""),
                "Safety stock": ltrow.get("Safety stock", ""),
                "Comp To date": ctd or "",
                "Notes": notes_val,
            }
        )

    amo_rows.sort(key=lambda r: (-r["Lead time"], r["Part number"]))
    return pd.DataFrame(
        amo_rows,
        columns=[
            "Part number",
            "Qty",
            "Lead time",
            "Item name",
            "Supplier",
            "Component name",
            "Component description",
            "Safety stock",
            "Comp To date",
            "Notes",
        ],
    )


def main():
    allbom_excel = "allbom.xlsx"
    output_xlsx = "AMO_Output_AllSheets.xlsx"

    edges, parents_set, comp_details, product_comp_to_date = load_component_db("component_db.csv")
    lead_map = load_leadtime("leadtime.csv")

    xf = pd.ExcelFile(allbom_excel, engine="openpyxl")
    sheet_names = xf.sheet_names

    used = set()
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for s in sheet_names:
            raw_df = pd.read_excel(xf, sheet_name=s, header=None)
            allbom_df = clean_allbom_sheet_df(raw_df)
            amo_df = process_allbom_df(allbom_df, edges, parents_set, comp_details, product_comp_to_date, lead_map)

            base = safe_excel_sheet_name(s)
            amo_sheet = safe_excel_sheet_name(f"AMO_{base}")

            if amo_sheet.lower() in used:
                i = 1
                orig = amo_sheet
                while amo_sheet.lower() in used:
                    suffix = f"_{i}"
                    amo_sheet = safe_excel_sheet_name(orig[: 31 - len(suffix)] + suffix)
                    i += 1

            amo_df.to_excel(writer, sheet_name=amo_sheet, index=False)
            used.add(amo_sheet.lower())

    print("Done! Created:", output_xlsx)
    print("Processed sheets:", len(sheet_names), sheet_names)

    from openpyxl import load_workbook

    wb = load_workbook(output_xlsx)
    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 1:
            continue
        headers = [cell.value for cell in ws[1]]
        if "Safety stock" not in headers:
            continue

        safety_col_idx = headers.index("Safety stock") + 1
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_filter_column(safety_col_idx - 1, ["0", "0.0", "0.00", 0])

    wb.save(output_xlsx)
    print("Applied filter Safety stock = 0 to all AMO sheets.")


if __name__ == "__main__":
    main()
