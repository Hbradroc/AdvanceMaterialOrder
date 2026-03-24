import csv
import re
from collections import defaultdict
from decimal import Decimal

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


THRESHOLD = 40
COMP_TO_DATE_REQUIRED = "99999999"
NOTES_PREFIXES = ("BGGXHWP", "BGGXKWP")
LOW_SAFETY_STOCK_THRESHOLD = 1.0

ALLBOM_PARTNO_CANDIDATES = [
    "Product number",
    "Product Number",
    "Product no",
    "Product No",
    "Product No.",
    "Component/WC",
    "Component/WC ",
    "Compenent/WC",
    "Compenent/WC ",
]

ALLBOM_QTY_CANDIDATES = [
    "Amount",
    "Qty",
    "Quantity",
    "quantity",
    "amount",
    "Qty/Min",
    "Qty / Min",
    "Qty/ Min",
    "Qty/Min ",
]

PRODUCT_CODE_COL_CANDIDATES = [
    "Product code",
    "Product Code",
    "Product",
    "Product name",
    "Unit",
    "Unit code",
    "Product ID",
    "Product No",
    "Product no",
    "Project",
    "Project code",
]


def normalize_id(x):
    s = (x or "").strip()
    if s == "":
        return ""
    low = s.casefold().strip()
    if low in {"n/a", "none", "nan", "na", ".", "..", "...", "-", "--", "---", "\ufffd"}:
        return ""
    if re.fullmatch(r"[-\.\ufffd]+", s):
        return ""
    s = s.replace(",", "")
    if re.fullmatch(r"\d+\.(0+)", s):
        s = s.split(".")[0]
    if re.fullmatch(r"\d+(\.\d+)?[eE][+-]?\d+", s):
        try:
            s = str(int(Decimal(s)))
        except Exception:
            pass
    return s.strip()


def to_float_safe(x):
    try:
        s = (x or "").strip()
        if s == "":
            return 0.0
        return float(s.replace(",", ""))
    except Exception:
        return 0.0


def to_int_safe(x):
    try:
        s = (x or "").strip()
        if s == "":
            return None
        s = s.replace(",", "")
        m = re.search(r"-?\d+(\.\d+)?", s)
        if not m:
            return None
        return int(float(m.group(0)))
    except Exception:
        return None


def normalize_makebuy(v):
    v = (v or "").strip().casefold()
    if ("purch" in v) or (v in {"buy", "bought", "purchase"}):
        return "purchased"
    if ("manuf" in v) or ("mfg" in v) or ("make" in v):
        return "manufacturing"
    return v


def normalize_acq(v):
    v = (v or "").strip().casefold()
    if "purch" in v:
        return "purchased"
    if ("manuf" in v) or ("mfg" in v) or ("make" in v):
        return "manufacturing"
    return v


def prefer_nonempty(old, new):
    old = (old or "").strip()
    new = (new or "").strip()
    return new if (new and not old) else old


def normalize_comp_to_date(x):
    s = (x or "").strip().replace(",", "")
    if re.fullmatch(r"\d+\.(0+)", s):
        s = s.split(".")[0]
    return s


def safe_excel_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "_", name or "").strip()
    name = name[:31] if len(name) > 31 else name
    return name or "Sheet"


def clean_allbom_sheet_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()

    if df.shape[1] > 0:
        first_col = df.iloc[:, 0]
        if first_col.isna().all() or first_col.astype(str).str.strip().str.replace("nan", "").eq("").all():
            df = df.iloc[:, 1:]

    header_row_idx = None
    candidates_cf = {c.strip().casefold() for c in ALLBOM_PARTNO_CANDIDATES}

    for i in range(min(60, len(df))):
        row_vals = df.iloc[i].tolist()
        row_strs = [
            str(x).strip().casefold()
            for x in row_vals
            if str(x).strip() and str(x).strip().casefold() != "nan"
        ]
        if not row_strs:
            continue
        if any(s in candidates_cf for s in row_strs):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise Exception(
            "Could not find AllBOM header row. Expected a column like 'Product number' or 'Component/WC'."
        )

    header = df.iloc[header_row_idx].tolist()
    df = df.iloc[header_row_idx + 1 :].copy()
    df.columns = [
        ""
        if (h is None or (isinstance(h, float) and pd.isna(h)))
        else str(h).strip()
        for h in header
    ]

    df = df.dropna(how="all")

    drop_cols = []
    for c in df.columns:
        c_str = str(c).strip()
        c_cf = c_str.casefold()
        if c_str == "" or c_cf == "nan" or c_cf.startswith("unnamed"):
            drop_cols.append(c)
    if drop_cols:
        df = df.drop(columns=drop_cols)

    col_map_cf = {str(c).strip().casefold(): c for c in df.columns}
    found_part_col = None
    for cand in ALLBOM_PARTNO_CANDIDATES:
        key = cand.strip().casefold()
        if key in col_map_cf:
            found_part_col = col_map_cf[key]
            break
    if found_part_col is None:
        raise Exception("Header found, but could not locate the part-number column after cleaning.")
    if found_part_col != "Product number":
        df = df.rename(columns={found_part_col: "Product number"})

    return df


def find_product_code_column(df: pd.DataFrame):
    cols = list(df.columns)
    lower_map = {str(c).strip().casefold(): c for c in cols}

    for cand in PRODUCT_CODE_COL_CANDIDATES:
        key = cand.strip().casefold()
        if key in lower_map:
            return lower_map[key]

    for c in cols:
        cl = str(c).strip().casefold()
        if "product" in cl and "code" in cl:
            return c
    return None


def build_part_to_productcode_map(allbom_df: pd.DataFrame) -> dict:
    part_to_code = {}
    if "Product number" not in allbom_df.columns:
        return part_to_code
    code_col = find_product_code_column(allbom_df)
    if code_col is None:
        return part_to_code

    for _, row in allbom_df.iterrows():
        pn = normalize_id(str(row.get("Product number", "") or ""))
        if not pn:
            continue
        code = str(row.get(code_col, "") or "").strip()
        if pn not in part_to_code and code:
            part_to_code[pn] = code
    return part_to_code


def load_component_db(path="component_db.csv"):
    edges = defaultdict(list)
    parents_set = set()
    comp_details = {}
    product_comp_to_date = {}

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        required = [
            "Product no",
            "Component no",
            "Component name",
            "Component description",
            "Make/buy",
            "Quantity",
            "Comp To date",
        ]
        missing = [c for c in required if c not in fields]
        if missing:
            raise Exception(f"{path} missing columns: {missing}. Found: {fields}")

        for r in reader:
            parent = normalize_id(r.get("Product no"))
            child = normalize_id(r.get("Component no"))
            if not parent or not child:
                continue

            edge_qty = to_float_safe(r.get("Quantity"))
            if edge_qty <= 0:
                edge_qty = 1.0

            parents_set.add(parent)
            edges[parent].append((child, edge_qty))

            ctd_parent = normalize_comp_to_date(r.get("Comp To date"))
            if ctd_parent:
                oldp = product_comp_to_date.get(parent, "")
                if oldp == COMP_TO_DATE_REQUIRED:
                    product_comp_to_date[parent] = oldp
                elif ctd_parent == COMP_TO_DATE_REQUIRED:
                    product_comp_to_date[parent] = ctd_parent
                elif oldp == "":
                    product_comp_to_date[parent] = ctd_parent

            existing = comp_details.get(child, {})
            mb_new = normalize_makebuy(r.get("Make/buy"))
            mb_old = normalize_makebuy(existing.get("Make/buy"))
            mb_final = "purchased" if (mb_new == "purchased" or mb_old == "purchased") else (mb_old or mb_new)

            ctd_new = normalize_comp_to_date(r.get("Comp To date"))
            ctd_old = normalize_comp_to_date(existing.get("Comp To date", ""))
            ctd_final = (
                COMP_TO_DATE_REQUIRED
                if (ctd_new == COMP_TO_DATE_REQUIRED or ctd_old == COMP_TO_DATE_REQUIRED)
                else (ctd_old or ctd_new)
            )

            comp_details[child] = {
                "Component no": child,
                "Component name": prefer_nonempty(existing.get("Component name"), r.get("Component name")),
                "Component description": prefer_nonempty(
                    existing.get("Component description"),
                    r.get("Component description"),
                ),
                "Make/buy": mb_final,
                "Comp To date": ctd_final,
            }

    return edges, parents_set, comp_details, product_comp_to_date


def load_leadtime(path="leadtime.csv"):
    lead_map = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        required = ["Item no", "Item name", "Lead time"]
        missing = [c for c in required if c not in fields]
        if missing:
            raise Exception(f"{path} missing columns: {missing}. Found: {fields}")

        acq_col = next((c for c in ("Acquisition code", "Acquisition", "Acq code", "Acq. code") if c in fields), None)
        supplier_col = next(
            (
                c
                for c in (
                    "Supplier",
                    "Supplier name",
                    "Supplier Name",
                    "Vendor",
                    "Vendor name",
                    "Vendor Name",
                    "Supplier/Vendor",
                    "Supplier/Vendor name",
                    "Manufacturer",
                    "Manufacturer name",
                    "Vendor No",
                    "Supplier No",
                )
                if c in fields
            ),
            None,
        )
        safety_col = next(
            (
                c
                for c in (
                    "Safety stock",
                    "Safety Stock",
                    "SafetyStock",
                    "Safety stock qty",
                    "Safety Stock Qty",
                    "Safety stock quantity",
                    "Safety",
                    "Safetystock",
                    "SS",
                    "Min stock",
                    "Minimum stock",
                    "Min qty",
                    "Minimum qty",
                    "Reorder point",
                    "ROP",
                )
                if c in fields
            ),
            None,
        )

        for r in reader:
            item = normalize_id(r.get("Item no"))
            if not item:
                continue
            lead_map[item] = {
                "Item no": item,
                "Item name": (r.get("Item name") or "").strip(),
                "Lead time": to_int_safe(r.get("Lead time")),
                "Acq": normalize_acq(r.get(acq_col)) if acq_col else "",
                "Supplier": (r.get(supplier_col) or "").strip() if supplier_col else "",
                "Safety stock": to_float_safe(r.get(safety_col)) if safety_col else "",
            }
    return lead_map


def process_allbom_df(allbom_df, edges, parents_set, comp_details, product_comp_to_date, lead_map):
    cols = list(allbom_df.columns)
    if "Product number" not in cols:
        raise Exception(f"AllBOM sheet missing 'Product number'. Found: {cols}")

    qty_col = next((c for c in ALLBOM_QTY_CANDIDATES if c in cols), None)
    part_to_productcode = build_part_to_productcode_map(allbom_df)

    allbom_qty = defaultdict(float)
    for _, row in allbom_df.iterrows():
        pn = normalize_id(str(row.get("Product number", "") or ""))
        if not pn:
            continue
        qty = to_float_safe(str(row.get(qty_col, "") or "")) if qty_col else 1.0
        if qty <= 0:
            continue
        allbom_qty[pn] += qty

    start_kits = []
    direct_parts_qty = defaultdict(float)
    for pn, qty in allbom_qty.items():
        if pn in parents_set:
            start_kits.append((pn, qty))
        else:
            direct_parts_qty[pn] += qty

    leaf_qty = defaultdict(float)

    def dfs(parent, parent_qty, depth, path):
        if depth > 80 or parent in path:
            return
        children = edges.get(parent, [])
        if not children:
            leaf_qty[parent] += parent_qty
            return
        for child, edge_qty in children:
            child_total_qty = parent_qty * edge_qty
            if child in parents_set:
                dfs(child, child_total_qty, depth + 1, path | {parent})
            else:
                leaf_qty[child] += child_total_qty

    for kit, kit_qty in start_kits:
        dfs(kit, kit_qty, 0, set())

    for part, qty in direct_parts_qty.items():
        leaf_qty[part] += qty

    def is_purchased(part_no):
        lt = lead_map.get(part_no)
        if lt and lt.get("Acq"):
            return lt["Acq"] == "purchased"
        return normalize_makebuy(comp_details.get(part_no, {}).get("Make/buy", "")) == "purchased"

    def get_comp_to_date_for_part(part_no):
        ctd = normalize_comp_to_date(comp_details.get(part_no, {}).get("Comp To date", ""))
        if ctd:
            return ctd
        return normalize_comp_to_date(product_comp_to_date.get(part_no, ""))

    amo_rows = []
    for part, total_qty in leaf_qty.items():
        if total_qty <= 0:
            continue
        ltrow = lead_map.get(part)
        if not ltrow or ltrow.get("Lead time") is None:
            continue
        if not is_purchased(part):
            continue
        if ltrow["Lead time"] <= THRESHOLD:
            continue

        ctd = get_comp_to_date_for_part(part)
        if ctd and ctd != COMP_TO_DATE_REQUIRED:
            continue

        notes_val = part_to_productcode.get(part, "") if str(part).startswith(NOTES_PREFIXES) else ""
        comp = comp_details.get(part, {})

        amo_rows.append(
            {
                "Part number": part,
                "Qty": round(total_qty, 3),
                "Lead time": ltrow["Lead time"],
                "Item name": ltrow.get("Item name", ""),
                "Supplier": ltrow.get("Supplier", ""),
                "Component name": comp.get("Component name", ""),
                "Component description": comp.get("Component description", ""),
                "Safety stock": ltrow.get("Safety stock", ""),
                "Comp To date": ctd or "",
                "Notes": notes_val,
            }
        )

    amo_rows.sort(key=lambda r: (-r["Lead time"], r["Part number"]))
    return pd.DataFrame(
        amo_rows,
        columns=[
            "Part number",
            "Qty",
            "Lead time",
            "Item name",
            "Supplier",
            "Component name",
            "Component description",
            "Safety stock",
            "Comp To date",
            "Notes",
        ],
    )


def add_dashboard_sheet(wb):
    if "AMO_Dashboard" in wb.sheetnames:
        del wb["AMO_Dashboard"]
    ws = wb.create_sheet("AMO_Dashboard", 0)

    all_rows = []
    per_sheet = defaultdict(list)
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("AMO_") or sheet_name == "AMO_Dashboard":
            continue
        sh = wb[sheet_name]
        if sh.max_row < 2:
            continue
        headers = [sh.cell(row=1, column=i).value for i in range(1, sh.max_column + 1)]
        header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v is not None}
        req = [
            "Part number",
            "Qty",
            "Lead time",
            "Supplier",
            "Safety stock",
            "Comp To date",
            "Component name",
            "Component description",
        ]
        if not all(k in header_map for k in req):
            continue
        for r in range(2, sh.max_row + 1):
            part = sh.cell(r, header_map["Part number"]).value
            if part in (None, ""):
                continue
            qty = sh.cell(r, header_map["Qty"]).value or 0
            lead = sh.cell(r, header_map["Lead time"]).value or 0
            supplier = sh.cell(r, header_map["Supplier"]).value or "Unknown"
            safety = sh.cell(r, header_map["Safety stock"]).value
            ctd = sh.cell(r, header_map["Comp To date"]).value or ""
            comp_name = sh.cell(r, header_map["Component name"]).value or ""
            comp_desc = sh.cell(r, header_map["Component description"]).value or ""
            try:
                qty = float(qty)
            except Exception:
                qty = 0.0
            try:
                lead = int(float(lead))
            except Exception:
                lead = 0
            try:
                safety = float(safety) if safety not in (None, "") else 0.0
            except Exception:
                safety = 0.0
            all_rows.append(
                {
                    "sheet": sheet_name,
                    "part": str(part),
                    "qty": qty,
                    "lead": lead,
                    "supplier": str(supplier).strip() or "Unknown",
                    "safety": safety,
                    "ctd": str(ctd).strip(),
                    "component_name": str(comp_name).strip(),
                    "component_description": str(comp_desc).strip(),
                }
            )
            per_sheet[sheet_name].append(all_rows[-1])

    # Basic styling
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=13)
    bold_font = Font(bold=True)

    ws["A1"] = "AMO Dashboard"
    ws["A1"].font = white_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:F1")

    total_items = len(all_rows)
    total_qty = round(sum(x["qty"] for x in all_rows), 3)
    avg_lead = round((sum(x["lead"] for x in all_rows) / total_items), 1) if total_items else 0
    available_count = sum(1 for x in all_rows if x["safety"] > 0)
    critical_count = total_items - available_count

    ws["A3"] = "Total AMO items"
    ws["B3"] = total_items
    ws["A4"] = "Total quantity"
    ws["B4"] = total_qty
    ws["A5"] = "Average lead time (days)"
    ws["B5"] = avg_lead
    ws["D3"] = "Available (Safety stock > 0)"
    ws["E3"] = available_count
    ws["D4"] = "Critical (Safety stock = 0)"
    ws["E4"] = critical_count
    for c in ("A3", "A4", "A5", "D3", "D4"):
        ws[c].font = bold_font

    # 1) Per-sheet breakdown
    ws["A8"] = "Breakdown by AMO Sheet"
    ws["A8"].font = bold_font
    ws["A9"], ws["B9"], ws["C9"], ws["D9"] = "Sheet", "Items", "Total Qty", "Avg Lead"
    for c in ("A9", "B9", "C9", "D9"):
        ws[c].font = bold_font
        ws[c].fill = header_fill
    row = 10
    sheet_stats = []
    for sheet_name in sorted(per_sheet.keys()):
        rows = per_sheet[sheet_name]
        items = len(rows)
        qty_sum = round(sum(x["qty"] for x in rows), 3)
        avg_l = round(sum(x["lead"] for x in rows) / items, 1) if items else 0
        sheet_stats.append((sheet_name, items, qty_sum, avg_l))
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=items)
        ws.cell(row=row, column=3, value=qty_sum)
        ws.cell(row=row, column=4, value=avg_l)
        row += 1

    # 2) Missing and low safety stock item detail
    no_safety = [x for x in all_rows if x["safety"] <= 0]
    low_safety = [x for x in all_rows if 0 < x["safety"] <= LOW_SAFETY_STOCK_THRESHOLD]

    ws["A18"] = "No Safety Stock Items (<= 0)"
    ws["A18"].font = bold_font
    ws["A19"], ws["B19"], ws["C19"], ws["D19"], ws["E19"], ws["F19"] = (
        "Part number",
        "Component name",
        "Component description",
        "Supplier",
        "Sheet",
        "Lead time",
    )
    for c in ("A19", "B19", "C19", "D19", "E19", "F19"):
        ws[c].font = bold_font
        ws[c].fill = header_fill
    row = 20
    for x in sorted(no_safety, key=lambda r: (-r["lead"], r["part"]))[:50]:
        ws.cell(row=row, column=1, value=x["part"])
        ws.cell(row=row, column=2, value=x.get("component_name", ""))
        ws.cell(row=row, column=3, value=x.get("component_description", ""))
        ws.cell(row=row, column=4, value=x["supplier"])
        ws.cell(row=row, column=5, value=x["sheet"])
        ws.cell(row=row, column=6, value=x["lead"])
        row += 1

    ws["A73"] = f"Low Safety Stock Items (<= {LOW_SAFETY_STOCK_THRESHOLD})"
    ws["A73"].font = bold_font
    ws["A74"], ws["B74"], ws["C74"], ws["D74"], ws["E74"], ws["F74"], ws["G74"] = (
        "Part number",
        "Safety stock",
        "Component name",
        "Component description",
        "Supplier",
        "Sheet",
        "Lead time",
    )
    for c in ("A74", "B74", "C74", "D74", "E74", "F74", "G74"):
        ws[c].font = bold_font
        ws[c].fill = header_fill
    row = 75
    for x in sorted(low_safety, key=lambda r: (r["safety"], -r["lead"], r["part"]))[:50]:
        ws.cell(row=row, column=1, value=x["part"])
        ws.cell(row=row, column=2, value=x["safety"])
        ws.cell(row=row, column=3, value=x.get("component_name", ""))
        ws.cell(row=row, column=4, value=x.get("component_description", ""))
        ws.cell(row=row, column=5, value=x["supplier"])
        ws.cell(row=row, column=6, value=x["sheet"])
        ws.cell(row=row, column=7, value=x["lead"])
        row += 1

    # 3) Longest lead-time items
    longest_lead_items = sorted(all_rows, key=lambda r: (-r["lead"], r["safety"], r["qty"], r["part"]))[:20]
    ws["A128"] = "Longest Lead Time Items"
    ws["A128"].font = bold_font
    ws["A129"], ws["B129"], ws["C129"], ws["D129"], ws["E129"] = (
        "Part number",
        "Lead time",
        "Qty",
        "Safety stock",
        "Supplier",
    )
    for c in ("A129", "B129", "C129", "D129", "E129"):
        ws[c].font = bold_font
        ws[c].fill = header_fill
    row = 130
    for x in longest_lead_items:
        ws.cell(row=row, column=1, value=x["part"])
        ws.cell(row=row, column=2, value=x["lead"])
        ws.cell(row=row, column=3, value=round(x["qty"], 3))
        ws.cell(row=row, column=4, value=x["safety"])
        ws.cell(row=row, column=5, value=x["supplier"])
        row += 1

    # Graph 1: qty by sheet
    if sheet_stats:
        bar_sheet = BarChart()
        bar_sheet.title = "Qty by AMO Sheet"
        bar_sheet.y_axis.title = "Qty"
        bar_sheet.x_axis.title = "Sheet"
        bar_sheet.type = "col"
        bar_sheet.style = 11
        max_row = 9 + len(sheet_stats)
        data_sheet = Reference(ws, min_col=3, min_row=9, max_row=max_row)
        cats_sheet = Reference(ws, min_col=1, min_row=10, max_row=max_row)
        bar_sheet.add_data(data_sheet, titles_from_data=True)
        bar_sheet.set_categories(cats_sheet)
        bar_sheet.height = 6
        bar_sheet.width = 9
        ws.add_chart(bar_sheet, "L3")

    # 4) Supplier quantities for low-priority-risk items only
    supplier_low_qty = defaultdict(float)
    for x in all_rows:
        if x["safety"] <= LOW_SAFETY_STOCK_THRESHOLD:
            supplier_low_qty[x["supplier"]] += x["qty"]
    low_suppliers = sorted(supplier_low_qty.items(), key=lambda kv: kv[1], reverse=True)[:10]
    ws["A153"] = "Supplier Quantities (Low Safety only)"
    ws["A153"].font = bold_font
    ws["A154"], ws["B154"] = "Supplier", "Qty"
    ws["A154"].font = ws["B154"].font = bold_font
    ws["A154"].fill = ws["B154"].fill = header_fill
    row = 155
    for s, q in low_suppliers:
        ws.cell(row=row, column=1, value=s)
        ws.cell(row=row, column=2, value=round(q, 3))
        row += 1

    # Graph-only area on the right side
    if longest_lead_items:
        bar2 = BarChart()
        bar2.title = "Longest Lead Time Items"
        bar2.y_axis.title = "Lead time"
        bar2.x_axis.title = "Part number"
        bar2.type = "bar"
        bar2.style = 10
        max_row = 129 + len(longest_lead_items)
        data2 = Reference(ws, min_col=2, min_row=129, max_row=max_row)
        cats2 = Reference(ws, min_col=1, min_row=130, max_row=max_row)
        bar2.add_data(data2, titles_from_data=True)
        bar2.set_categories(cats2)
        bar2.height = 7
        bar2.width = 9
        ws.add_chart(bar2, "L21")

    ws["D153"] = "Availability"
    ws["D153"].font = bold_font
    ws["D154"], ws["E154"] = "Status", "Count"
    ws["D154"].font = ws["E154"].font = bold_font
    ws["D154"].fill = ws["E154"].fill = header_fill
    ws["D155"], ws["E155"] = "Available", available_count
    ws["D156"], ws["E156"] = "Critical", critical_count
    pie = PieChart()
    pie.title = "Available vs Critical Items"
    pie_data = Reference(ws, min_col=5, min_row=154, max_row=156)
    pie_labels = Reference(ws, min_col=4, min_row=155, max_row=156)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_labels)
    pie.height = 6
    pie.width = 8
    pie.style = 10
    ws.add_chart(pie, "L39")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 22


def main():
    allbom_excel = "allbom.xlsx"
    output_xlsx = "AMO_Output_AllSheets.xlsx"

    edges, parents_set, comp_details, product_comp_to_date = load_component_db("component_db.csv")
    lead_map = load_leadtime("leadtime.csv")

    xf = pd.ExcelFile(allbom_excel, engine="openpyxl")
    sheet_names = xf.sheet_names

    used = set()
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for s in sheet_names:
            raw_df = pd.read_excel(xf, sheet_name=s, header=None)
            allbom_df = clean_allbom_sheet_df(raw_df)
            amo_df = process_allbom_df(allbom_df, edges, parents_set, comp_details, product_comp_to_date, lead_map)

            base = safe_excel_sheet_name(s)
            amo_sheet = safe_excel_sheet_name(f"AMO_{base}")

            if amo_sheet.lower() in used:
                i = 1
                orig = amo_sheet
                while amo_sheet.lower() in used:
                    suffix = f"_{i}"
                    amo_sheet = safe_excel_sheet_name(orig[: 31 - len(suffix)] + suffix)
                    i += 1

            amo_df.to_excel(writer, sheet_name=amo_sheet, index=False)
            used.add(amo_sheet.lower())

    print("Done! Created:", output_xlsx)
    print("Processed sheets:", len(sheet_names), sheet_names)

    from openpyxl import load_workbook

    wb = load_workbook(output_xlsx)
    for ws in wb.worksheets:
        if ws.max_row < 2 or ws.max_column < 1:
            continue
        headers = [cell.value for cell in ws[1]]
        if "Safety stock" not in headers:
            continue

        safety_col_idx = headers.index("Safety stock") + 1
        ws.auto_filter.ref = ws.dimensions
        ws.auto_filter.add_filter_column(safety_col_idx - 1, ["0", "0.0", "0.00", 0])

    add_dashboard_sheet(wb)

    wb.save(output_xlsx)
    print("Applied filter Safety stock = 0 to all AMO sheets.")
    print("Added AMO_Dashboard with charts.")


if __name__ == "__main__":
    main()
